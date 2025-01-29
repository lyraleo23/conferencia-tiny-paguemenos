import os
import asyncio
import time
import pandas as pd
import openpyxl
from dotenv import load_dotenv
from api_tiny_v3 import obter_pedidos_v3, obter_marcadores_v3, obter_pedido_v3, obter_nota_fiscal_v3
from api_miliapp import obter_tokens_tiny

load_dotenv()
TOKEN_TINY = os.getenv('TOKEN_TINY')
TOKEN_MILIAPP = os.getenv('TOKEN_MILIAPP')
ACCESS_TOKEN, REFRESH_TOKEN = obter_tokens_tiny(TOKEN_MILIAPP, 'miligrama')
ACCESS_TOKEN_FOR, REFRESH_TOKEN_FOR = obter_tokens_tiny(TOKEN_MILIAPP, 'miligrama_nordeste')

async def main():
    os.system('cls')

    # Carrega a planilha
    workbook = openpyxl.load_workbook("Pedidos.xlsx")

    new_xlsx = [
        ['marketplace', 'numero_marketplace', 'numero_ecommerce', 'numero_pedido', 'situacao', 'conta_tiny', 'numero_nota_fiscal', 'serie_nota_fiscal', 'chave_nota_fiscal', 'data_nota_fiscal', 'hora_emissao']
    ]

    # Verifica quantas linhas tem na planilha
    sheet = workbook.active
    num_linhas = sheet.max_row

    # Itera sobre as linhas da planilha
    for row in sheet.iter_rows(min_row=2, max_row=num_linhas, values_only=True):
        while True:
            try:
                print(row[4])

                marketplace = row[1]
                print(marketplace)
                numero_marketplace = row[5]
                numero_ecommerce = None

                # define o código do marketplace
                if marketplace == 'PAGUE MENOS':
                    numero_ecommerce = f'PGM-{row[5]}'
                print(numero_ecommerce)

                # Buscar o pedido na tiny
                pedidos_tiny = buscar_pedido_marketplace(numero_ecommerce)
                time.sleep(1)

                transferido_nordeste = False
                conta_tiny = 'Miligrama'

                if len(pedidos_tiny) > 0:
                    for p in pedidos_tiny:
                        id_pedido = p['id']
                        situacao = p['situacao']
                        numero_pedido = p['numeroPedido']

                        if situacao == 2:
                            transferido_nordeste = verifica_transferencia(p)
                            print(f'transferido_nordeste: {transferido_nordeste}')
                            
                            if transferido_nordeste == True:
                                pedido_detalhado = procurar_pedido_fortaleza(p)
                                id_pedido = pedido_detalhado['id']
                                situacao = pedido_detalhado['situacao']
                                numero_pedido = pedido_detalhado['numeroPedido']
                                id_nota_fiscal = pedido_detalhado['idNotaFiscal']
                                conta_tiny = 'Miligrama Nordeste'

                                dados_nota_fiscal = obter_nota_fiscal_v3(ACCESS_TOKEN_FOR, id_nota_fiscal)
                                time.sleep(1)

                        situacao_nome = obter_nome_situacao(situacao)

                        if conta_tiny == 'Miligrama':
                            pedido_detalhado = obter_pedido_v3(ACCESS_TOKEN, id_pedido)
                            time.sleep(1)
                            id_nota_fiscal = pedido_detalhado['idNotaFiscal']
                            dados_nota_fiscal = obter_nota_fiscal_v3(ACCESS_TOKEN, id_nota_fiscal)
                            time.sleep(1)
                        
                        numero_nota_fiscal = dados_nota_fiscal['numero']
                        serie_nota_fiscal = dados_nota_fiscal['serie']
                        data_nota_fiscal = dados_nota_fiscal['dataEmissao']
                        data_nota_fiscal = data_nota_fiscal.split('-')
                        data_nota_fiscal = f'{data_nota_fiscal[2]}/{data_nota_fiscal[1]}/{data_nota_fiscal[0]}'
                        hora_nota_fiscal = dados_nota_fiscal['dataInclusao']
                        hora_nota_fiscal = hora_nota_fiscal.split()
                        hora_nota_fiscal = hora_nota_fiscal[1]
                        chave_nota_fiscal = dados_nota_fiscal['chaveAcesso']
                        
                        new_line = [marketplace, numero_marketplace, numero_ecommerce, numero_pedido, situacao, conta_tiny, numero_nota_fiscal, serie_nota_fiscal, chave_nota_fiscal, data_nota_fiscal, hora_nota_fiscal]
                        new_xlsx.append(new_line)
                break
            except:
                print('Tentando novamente...')

        # Save the new_xlsx data to an Excel file
        df = pd.DataFrame(new_xlsx)
        df.to_excel('conferencia.xlsx', index=True, header=True)
                
def buscar_pedido_marketplace(numero_ecommerce):
    params = {
        'numeroPedidoEcommerce': numero_ecommerce
    }
    return obter_pedidos_v3(ACCESS_TOKEN, params)

def obter_nome_situacao(situacao):
    match situacao:
        case 0:
            situacao_nome = 'aberto'
        case 1:
            situacao_nome = 'faturado'
        case 2:
            situacao_nome = 'cancelado'
        case 3:
            situacao_nome = 'aprovado'
        case 4:
            situacao_nome = 'preparando_envio'
        case 5:
            situacao_nome = 'enviado'
        case 6:
            situacao_nome = 'entregue'
        case 7:
            situacao_nome = 'pronto_envio'
        case 8:
            situacao_nome = 'dados_incompletos'
        case 9:
            situacao_nome = 'nao_entregue'
    
    return situacao_nome

def verifica_transferencia(pedidoTiny):
    time.sleep(1)
    transferido = False
    while True:
        try:
            marcadores = obter_marcadores_v3(ACCESS_TOKEN, pedidoTiny['id'])

            if marcadores:
                for marcador in marcadores:
                    if marcador['descricao'] == 'transferido multiempresa':
                        transferido = True
            break
        except:
            time.sleep(1)
            continue

    time.sleep(1)
    return transferido

def procurar_pedido_fortaleza(pedidoTiny):
    time.sleep(1)
    print('Procurando pedido em Fortaleza')
    print(pedidoTiny['ecommerce']['numeroPedidoEcommerce'])
    cpfCnpj = pedidoTiny['cliente']['cpfCnpj']
    params = {
        'cpfCnpj': cpfCnpj
    }
    lista_pedidos_cliente = obter_pedidos_v3(ACCESS_TOKEN_FOR, params)
    time.sleep(1)

    for pedido in lista_pedidos_cliente:
        id_pedido = pedido['id']
        pedido_detalhado = obter_pedido_v3(ACCESS_TOKEN_FOR, id_pedido)
        time.sleep(1)
        try:
            print(pedido_detalhado['numeroOrdemCompra'])
            if pedido_detalhado['numeroOrdemCompra'] == pedidoTiny['ecommerce']['numeroPedidoEcommerce']:
                print('Pedido encontrado')
                return pedido_detalhado
        except:
            continue

    return

asyncio.run(main())
